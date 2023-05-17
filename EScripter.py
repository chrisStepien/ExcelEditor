from openpyxl import load_workbook
import os

def main():
    
    while True:    
        path = get_file()
        print(path)
        wb = load_workbook(path)
        
        menu(wb, path)
        
def menu(workbook, path):
    while True:
        options = {"F": "Filter Worksheet", "X": "Exit"}
        
        #Add loop for option dict when adding another option
        print("************************************")
        print("*           EDITOR MENU            *")    
        print("************************************")    
        print("| OPTIONS:                         |")
        print("|----------------------------------|")
        print("| F) Filter Worksheet              |")
        print("| ...                              |")
        print("| X) Exit                          |")
        print("|__________________________________|")
        print("")    
            
        
        try:
            user_input = str(input("Please select an option: "))
            
        except ValueError:
            print("Sorry, I didn't understand that.")        
        
        for key in options:
            if user_input.lower() == options[key].lower() or user_input.upper() == key:
                selection = key
                if selection.upper() == 'F':
                    filter(workbook, path)
                elif selection.upper() == 'X':
                    return                   
        
def filter(wb, path):
    
    while True:
        #Worksheet Menu
        print("")    
        print("-----------------------------")
        print("| Worksheets:               |")
        print("-----------------------------")
        for idx, sheet in enumerate(wb.sheetnames):
                
                dataset = "| " + str(idx + 1) + ": " + str(sheet)
                offset = 28 - len(dataset)
                dataset += ' ' * offset + '|'
                
                print(dataset)
        print("| X: Back                   |")            
        print("|___________________________|")
        print("") 
        
        try:
            user_input = str(input("Which worksheet would you like to filter: "))
        
        except ValueError:
            print("Sorry, I didn't understand that.")        
        
        
        if user_input.isnumeric() and int(user_input) <= len(wb.sheetnames):
            sheet = wb.sheetnames[int(user_input) - 1]
            ws = wb[sheet]
        elif user_input in wb.sheetnames:
            sheet = user_input
            ws = wb[sheet]
        elif user_input.upper() == 'X':
            return
        else:
            ws = None
            print("Sorry, I didn't understand that.")        
        
        if ws:      
                # for row in ws.iter_rows():
                #     print(row[0])
            while True:    
    
                #Column Header Menu
                print("")    
                print("-----------------------------")
                print("| Headers:                  |")
                print("-----------------------------")
                for idx, col in enumerate(ws.iter_cols()):
                    
                    dataset = "| " + str(idx + 1) + ": " + str(col[0].value)
                    offset = 28 - len(dataset)
                    dataset += ' ' * offset + '|'
                    
                    print(dataset)
                
                print("| X: Back                   |")    
                print("|___________________________|")
                print("")
                    
                try:
                    user_input = str(input("Which column would you like to filter with: "))
            
                except ValueError:
                    print("Sorry, I didn't understand that.")
                
                if user_input.isnumeric() and int(user_input) <= ws.max_column:
                    selection = ws[1][int(user_input) - 1]
                    print(selection)
                    condition = str(input("Enter filter value: "))
                    new_worksheet = str(input("Enter a name for your new worksheet: "))
                    
                    wb.create_sheet(new_worksheet)
                    ws2 = wb[new_worksheet]
                    
                    #copy_header(wb, ws, ws2)
                    
                    print("*----Worksheet Created----*")
                    
                    for row in ws.iter_rows():
                        if row[int(user_input) - 1].value == condition:
                            ws2.append((cell.value for cell in row))                         
                    
                    wb.save(path)
                    print("*----Workbook Saved----*")
                    print("")
                    
                    return
                #elif for String input needed
                
                elif user_input.upper() == 'X':
                    return
                else:
                    print("Sorry, I didn't understand that.") 

# def copy_header(workbook, src, dst):
    
#     num = 


def get_file():
    
    allowed_types = ['.xlsx']
    close_commands = ['EXIT', 'E', 'X', 'QUIT', 'Q']
    
    while True:
        try:
            file_name = str(input("Please enter file name: "))
        except ValueError:
            print("Sorry, I didn't understand that.")
        
        if file_name:
            file_path = "./Workbooks/" + file_name
            for type in allowed_types:
                if os.path.exists(file_path):
                    full_path = file_path
                    return full_path
                
                if os.path.exists(file_path + type):
                    full_path = file_path + type
                    return full_path
            
            if file_name.upper() in close_commands:
                quit()        
            
if __name__ == "__main__":
    main()        